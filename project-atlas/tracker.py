from openpyxl import Workbook, load_workbook
from datetime import datetime
import os


FILE = "data/applied.xlsx"




def log(job, status , note=""):
    os.makedirs("data", exist_ok=True)

    if not os.path.exists(FILE):
        wb = Workbook()
        ws = wb.active
        ws.append([ "Company", "Role", "Location", "Platform","Status","Note","Date", "Time",  "Link"])
        wb.save(FILE)

    wb = load_workbook(FILE)
    ws = wb.active

    now = datetime.now()

    ws.append([
        job.get("company"),
        job.get("title"),
        job.get("location"),
        job.get("platform"),
        status,
        note,
        now.strftime("%d-%m-%y"),      # Date → 19-01-26
        now.strftime("%H:%M"),   # Time → 18:30
        job.get("link")
    ])

    wb.save(FILE)